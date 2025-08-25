import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import load_workbook

from app.utils.patterns import extract_job_numbers, extract_revisions

logger = logging.getLogger(__name__)


class ExcelParser:
    """Service for parsing Excel BOM files"""
    
    def __init__(self):
        self.logger = logger
    
    def read_excel_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read Excel file and return all sheets as DataFrames"""
        try:
            # Try to read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    sheets[sheet_name] = df
                except Exception as e:
                    self.logger.warning(f'Failed to read sheet {sheet_name}: {str(e)}')
                    continue
            
            return sheets
            
        except Exception as e:
            self.logger.error(f'Failed to read Excel file {file_path}: {str(e)}')
            return {}
    
    def find_bom_data_in_sheet(self, df: pd.DataFrame) -> Optional[Dict[str, Any]]:
        """Find BOM data in a worksheet by looking for expected column patterns"""
        try:
            # Look for rows that might contain headers
            potential_header_rows = []
            
            for idx, row in df.iterrows():
                row_text = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
                
                # Look for BOM-like headers
                if any(keyword in row_text for keyword in ['job', 'part', 'revision', 'qty', 'description']):
                    potential_header_rows.append(idx)
            
            if not potential_header_rows:
                return None
            
            # Use the first potential header row
            header_row = potential_header_rows[0]
            
            # Create new DataFrame starting from header row
            bom_df = df.iloc[header_row:].copy()
            bom_df.columns = bom_df.iloc[0]  # Use first row as headers
            bom_df = bom_df.drop(bom_df.index[0])  # Remove header row
            
            # Clean column names
            bom_df.columns = [str(col).strip() if pd.notna(col) else f'Column_{i}' 
                             for i, col in enumerate(bom_df.columns)]
            
            return {
                'dataframe': bom_df,
                'header_row': header_row,
                'columns': list(bom_df.columns)
            }
            
        except Exception as e:
            self.logger.error(f'Failed to find BOM data in sheet: {str(e)}')
            return None
    
    def extract_bom_data_standard_format(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract BOM data assuming standard format (Column A: Job, B: Part, H: Revision)"""
        try:
            # Map standard columns
            column_mapping = {
                'job_numbers': 0,    # Column A
                'part_numbers': 1,   # Column B  
                'revisions': 7       # Column H
            }
            
            extracted_data = {
                'job_numbers': [],
                'part_numbers': [],
                'revisions': []
            }
            
            # Extract data from each column
            for data_type, col_index in column_mapping.items():
                if col_index < len(df.columns):
                    column_data = df.iloc[:, col_index].dropna()
                    
                    for value in column_data:
                        str_value = str(value).strip()
                        if str_value and str_value.lower() not in ['nan', 'none', '']:
                            extracted_data[data_type].append(str_value)
            
            return extracted_data
            
        except Exception as e:
            self.logger.error(f'Failed to extract standard format BOM data: {str(e)}')
            return {'job_numbers': [], 'part_numbers': [], 'revisions': []}
    
    def extract_bom_data_flexible(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract BOM data with flexible column detection"""
        try:
            extracted_data = {
                'job_numbers': [],
                'part_numbers': [],
                'revisions': []
            }
            
            # Search all cells for patterns
            for _, row in df.iterrows():
                for cell in row:
                    if pd.isna(cell):
                        continue
                    
                    cell_text = str(cell).strip()
                    
                    # Extract job numbers (5 digits)
                    job_nums = extract_job_numbers(cell_text)
                    extracted_data['job_numbers'].extend(job_nums)
                    
                    # Extract revisions
                    revisions = extract_revisions(cell_text)
                    extracted_data['revisions'].extend(revisions)
                    
                    # Look for part number patterns
                    if any(prefix in cell_text.upper() for prefix in ['PCA-', 'DRW-', 'PCB-']):
                        extracted_data['part_numbers'].append(cell_text)
            
            # Remove duplicates
            for key in extracted_data:
                extracted_data[key] = list(set(extracted_data[key]))
            
            return extracted_data
            
        except Exception as e:
            self.logger.error(f'Failed to extract flexible BOM data: {str(e)}')
            return {'job_numbers': [], 'part_numbers': [], 'revisions': []}
    
    def parse_bom_file(self, file_path: str) -> Dict[str, Any]:
        """Parse Excel BOM file and extract relevant data"""
        try:
            # Read all sheets
            sheets = self.read_excel_file(file_path)
            
            if not sheets:
                return {
                    'success': False,
                    'error': 'No readable sheets found in Excel file',
                    'file_type': 'BOM_EXCEL'
                }
            
            all_extracted_data = {
                'job_numbers': [],
                'part_numbers': [],
                'revisions': []
            }
            
            sheet_results = {}
            
            # Process each sheet
            for sheet_name, df in sheets.items():
                self.logger.info(f'Processing sheet: {sheet_name}')
                
                # Try to find structured BOM data
                bom_info = self.find_bom_data_in_sheet(df)
                
                if bom_info:
                    # Extract data using standard format
                    standard_data = self.extract_bom_data_standard_format(bom_info['dataframe'])
                else:
                    # Fallback to flexible extraction
                    standard_data = self.extract_bom_data_flexible(df)
                
                # Merge data
                for key in all_extracted_data:
                    all_extracted_data[key].extend(standard_data[key])
                
                sheet_results[sheet_name] = {
                    'has_bom_structure': bom_info is not None,
                    'extracted_data': standard_data,
                    'row_count': len(df),
                    'column_count': len(df.columns)
                }
            
            # Remove duplicates from combined data
            for key in all_extracted_data:
                all_extracted_data[key] = list(set(all_extracted_data[key]))
            
            result = {
                'success': True,
                'file_type': 'BOM_EXCEL',
                'extraction_method': 'pandas+openpyxl',
                'sheet_count': len(sheets),
                'sheet_results': sheet_results,
                **all_extracted_data
            }
            
            self.logger.info(f'Successfully parsed BOM file: {file_path}')
            return result
            
        except Exception as e:
            self.logger.error(f'BOM file parsing failed for {file_path}: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'file_type': 'BOM_EXCEL'
            }
    
    def get_excel_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get basic information about the Excel file"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                }
            
            return {
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'sheet_info': sheet_info
            }
            
        except Exception as e:
            self.logger.error(f'Failed to get Excel file info for {file_path}: {str(e)}')
            return {}
    
    def validate_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Validate that the Excel file is readable and contains data"""
        if not Path(file_path).exists():
            return {
                'valid': False,
                'error': 'File does not exist'
            }
        
        try:
            # Try to open with openpyxl
            workbook = load_workbook(file_path, read_only=True)
            
            if not workbook.sheetnames:
                return {
                    'valid': False,
                    'error': 'Excel file contains no sheets'
                }
            
            # Check if sheets have data
            has_data = False
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.max_row > 1 and sheet.max_column > 1:
                    has_data = True
                    break
            
            file_info = self.get_excel_file_info(file_path)
            
            return {
                'valid': True,
                'has_data': has_data,
                **file_info
            }
            
        except Exception as e:
            return {
                'valid': False,
                'error': f'Excel file validation failed: {str(e)}'
            }
